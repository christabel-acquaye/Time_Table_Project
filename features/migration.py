import openpyxl
from os import path
import pandas as pd
import pprint
from features.exam.services import insert_exam
from features.rooms.services import insert_rooms
from features.periods.services import insert_period

# Function that reads the details from the excel file and inserts them into the exams table
def read_exam(insert = False):
    file_path = path.join(path.dirname(path.abspath(__file__)), '../data')
    # print(file_path)
    book = openpyxl.load_workbook(file_path + '/exam_input_data.xlsx')
    exam_data = pd.read_excel(file_path + '/exam_input_data.xlsx', sheet_name='exams')
    print(exam_data.to_dict('record'))
    # print(exam_data.to_dict('record'))

    normalized_data = []

    for data in exam_data.to_dict('record'):
        item = {
            'id' : data['ExamID'],
            'length': data['Length'] ,
            'alt': data['Alt_Seating'],
            'minSize' : data['MinSize'],
            'maxRooms' : data['MaxRooms'],
            'average' : data['Average'],
            'examCode' : data['ExamCode']
        }
        normalized_data.append(item)
        if insert:
            insert_exam(**item)
        # print(data)

    return normalized_data

# Function that reads the details from the excel file and inserts them into the room table
def read_room(insert = False):
    file_path = path.join(path.dirname(path.abspath(__file__)), '../data')
    # print(file_path)
    book = openpyxl.load_workbook(file_path + '/exam_input_data.xlsx')
    room_data = pd.read_excel(file_path + '/exam_input_data.xlsx', sheet_name='rooms')
    # print(room_data.to_dict('record'))

    normalized_data = []
    for data in room_data.to_dict('record'):


        item = {
            'id' : data['RoomID'],
            'roomName': data['RoomName'] ,
            'alt': data['Alt_Size'],
            'size' : data['Size'],
            'coord_longitude' : data['Coord_Longitude'],
            'coord_latitude' : data['Coord_Latitude'],

        }
        normalized_data.append(item)
        if insert:
            insert_rooms(**item)
        # print(data)

    return normalized_data


# Function that reads the details from the excel file and inserts them into the periods table
def read_period(insert = False):
    file_path = path.join(path.dirname(path.abspath(__file__)), '../data')
    # print(file_path)
    book = openpyxl.load_workbook(file_path + '/exam_input_data.xlsx')
    period_data = pd.read_excel(file_path + '/exam_input_data.xlsx', sheet_name='periods')
    print(period_data.to_dict('record'))

    normalized_data = []

    for data in period_data.to_dict('record'):
        item = {
            'id' : data['PeriodID'],
            'length': data['Length'] ,
            'day': data['Date'],
            'time' : data['Time'],
            'penalty' : data['Penalty']

        }
        normalized_data.append(item)
        if insert:
            insert_period(**item)
        # print(data)

    return normalized_data


# Function that reads the details from the excel file and organizes it in a more usable form
def read_student():
    file_path = path.join(path.dirname(path.abspath(__file__)), '../data')
    # print(file_path)
    book = openpyxl.load_workbook(file_path + '/exam_input_data.xlsx')
    student_data = pd.read_excel(file_path + '/exam_input_data.xlsx', sheet_name='students')
    
  
    student_group = []
    for data in student_data.to_dict('record'):
        item = list(data.values())
        dic = {
        'id' : item[0], 
        'course' : item[1:]
        }
       
        flipped = []

        for course in dic['course']:
            # pprint.pprint(course)
            if course not in flipped:
                flipped.append(course)
        dic['course'] = flipped
        student_group.append(dic)
        
    
    return student_group
        


       
if __name__ == '__main__':
    read_period(insert = True)